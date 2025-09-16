from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Dict, List, Union

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


HEADERS = [
    "id",
    "title_ar",
    "descriptionHtml_ar",
    "tags",
    "title_en",
    "descriptionHtml_en",
    "original_language",
    "target_language",
]


def autofit(ws) -> None:
    for col_idx, header in enumerate(HEADERS, start=1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            val = row[0].value
            if val is None:
                continue
            length = len(str(val))
            if length > max_len:
                max_len = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)


def main() -> None:
    if len(sys.argv) < 3:
        print("Usage:")
        print("  python scripts/json_to_xlsx.py <input_json> <output_xlsx>")
        print("\nExamples:")
        print("  python scripts/json_to_xlsx.py exports/single_product_with_lang.json exports/single_product_with_lang.xlsx")
        print("  python scripts/json_to_xlsx.py exports/products_with_lang.json exports/products_with_lang.xlsx")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    if not input_path.exists():
        print(f"Input not found: {input_path}")
        sys.exit(1)

    data = json.loads(input_path.read_text(encoding="utf-8"))

    rows: List[Dict[str, Union[str, int]]] = data if isinstance(data, list) else [data]

    wb = Workbook()
    ws = wb.active
    ws.title = "translations"

    # headers
    ws.append(HEADERS)

    for row in rows:
        # Convert tags list to string for Excel compatibility
        row_data = []
        for k in HEADERS:
            value = row.get(k, "")
            if k == "tags" and isinstance(value, list):
                value = ", ".join(value) if value else ""
            row_data.append(value)
        ws.append(row_data)

    autofit(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Wrote {len(rows)} rows -> {output_path}")


if __name__ == "__main__":
    main()
