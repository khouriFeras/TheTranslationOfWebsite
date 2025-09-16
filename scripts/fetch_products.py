#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fetch_products.py
- Unified script for fetching Shopify products
- Supports: all products, single product (by handle/id/tag), collection products
- Exports to JSON, CSV, and Excel formats
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import requests
from dotenv import load_dotenv

# Load environment variables from .env if present
load_dotenv()

SHOPIFY_STORE_DOMAIN: str = os.getenv("SHOPIFY_STORE_DOMAIN", "").strip()
SHOPIFY_ADMIN_ACCESS_TOKEN: str = os.getenv("SHOPIFY_ADMIN_ACCESS_TOKEN", "").strip()
SHOPIFY_API_VERSION: str = os.getenv("SHOPIFY_API_VERSION", "2024-07").strip()
TARGET_LANGUAGE: str = os.getenv("TARGET_LANGUAGE", "en").strip()
ORIGINAL_LANGUAGE: str = os.getenv("ORIGINAL_LANGUAGE", "ar").strip()
OUTPUT_DIR: str = os.getenv("OUTPUT_DIR", "exports").strip()


def validate_environment() -> None:
    missing: List[str] = []
    if not SHOPIFY_STORE_DOMAIN:
        missing.append("SHOPIFY_STORE_DOMAIN")
    if not SHOPIFY_ADMIN_ACCESS_TOKEN:
        missing.append("SHOPIFY_ADMIN_ACCESS_TOKEN")
    if missing:
        raise SystemExit(
            f"Missing required environment variables: {', '.join(missing)}.\n"
            "Set them in .env or your environment and try again."
        )


def graphql_endpoint() -> str:
    return f"https://{SHOPIFY_STORE_DOMAIN}/admin/api/{SHOPIFY_API_VERSION}/graphql.json"


def shopify_headers() -> Dict[str, str]:
    return {
        "X-Shopify-Access-Token": SHOPIFY_ADMIN_ACCESS_TOKEN,
        "Content-Type": "application/json",
    }


# GraphQL Queries
PRODUCTS_QUERY: str = """
query GetProducts($first: Int!, $after: String) {
  products(first: $first, after: $after) {
    pageInfo {
      hasNextPage
      endCursor
    }
    edges {
      node {
        id
        title
        handle
        descriptionHtml
        tags
        productType
        vendor
        createdAt
        updatedAt
        status
        publishedAt
      }
    }
  }
}
"""

PRODUCT_BY_HANDLE_QUERY = """
query GetProductByHandle($handle: String!) {
  product(handle: $handle) {
    id
    title
    handle
    descriptionHtml
    tags
    productType
    vendor
    createdAt
    updatedAt
    status
    publishedAt
  }
}
"""

PRODUCT_BY_ID_QUERY = """
query GetProductById($id: ID!) {
  product(id: $id) {
    id
    title
    handle
    descriptionHtml
    tags
    productType
    vendor
    createdAt
    updatedAt
    status
    publishedAt
  }
}
"""

PRODUCTS_BY_TAG_QUERY = """
query GetProductsByTag($tag: String!, $first: Int!, $after: String) {
  products(first: $first, after: $after, query: $tag) {
    pageInfo {
      hasNextPage
      endCursor
    }
    edges {
      node {
        id
        title
        handle
        descriptionHtml
        tags
        productType
        vendor
        createdAt
        updatedAt
        status
        publishedAt
      }
    }
  }
}
"""

COLLECTIONS_QUERY = """
query GetCollections($first: Int!, $after: String) {
  collections(first: $first, after: $after) {
    pageInfo {
      hasNextPage
      endCursor
    }
    edges {
      node {
        id
        title
        handle
      }
    }
  }
}
"""

COLLECTION_PRODUCTS_QUERY = """
query GetCollectionProducts($id: ID!, $first: Int!, $after: String) {
  collection(id: $id) {
    id
    title
    handle
    products(first: $first, after: $after) {
      pageInfo {
        hasNextPage
        endCursor
      }
      edges {
        node {
          id
          title
          handle
          descriptionHtml
          tags
          productType
          vendor
          createdAt
          updatedAt
          status
          publishedAt
        }
      }
    }
  }
}
"""


def make_graphql_request(query: str, variables: Optional[Dict] = None) -> Dict:
    """Make a GraphQL request to Shopify API."""
    if variables is None:
        variables = {}
    
    response = requests.post(
        graphql_endpoint(),
        headers=shopify_headers(),
        json={"query": query, "variables": variables}
    )
    
    if response.status_code != 200:
        raise SystemExit(f"HTTP {response.status_code}: {response.text}")
    
    data = response.json()
    if "errors" in data:
        raise SystemExit(f"GraphQL errors: {data['errors']}")
    
    return data["data"]


def fetch_all_products() -> List[Dict]:
    """Fetch all products from the store."""
    print("Fetching all products...")
    products = []
    has_next_page = True
    after = None
    page = 1
    
    while has_next_page:
        print(f"  Page {page}...")
        
        variables = {"first": 250, "after": after}
        data = make_graphql_request(PRODUCTS_QUERY, variables)
        
        products_data = data["products"]
        has_next_page = products_data["pageInfo"]["hasNextPage"]
        after = products_data["pageInfo"]["endCursor"]
        
        for edge in products_data["edges"]:
            products.append(edge["node"])
        
        page += 1
        
        # Rate limiting
        time.sleep(0.5)
    
    print(f"  Fetched {len(products)} products")
    return products


def fetch_product_by_handle(handle: str) -> Optional[Dict]:
    """Fetch a single product by handle."""
    print(f"Fetching product by handle: {handle}")
    
    variables = {"handle": handle}
    data = make_graphql_request(PRODUCT_BY_HANDLE_QUERY, variables)
    
    return data.get("product")


def fetch_product_by_id(product_id: str) -> Optional[Dict]:
    """Fetch a single product by ID."""
    print(f"Fetching product by ID: {product_id}")
    
    # Ensure ID is in correct format
    if not product_id.startswith("gid://shopify/Product/"):
        product_id = f"gid://shopify/Product/{product_id}"
    
    variables = {"id": product_id}
    data = make_graphql_request(PRODUCT_BY_ID_QUERY, variables)
    
    return data.get("product")


def fetch_products_by_tag(tag: str) -> List[Dict]:
    """Fetch products by tag."""
    print(f"Fetching products by tag: {tag}")
    products = []
    has_next_page = True
    after = None
    page = 1
    
    while has_next_page:
        print(f"  Page {page}...")
        
        variables = {"tag": f"tag:{tag}", "first": 250, "after": after}
        data = make_graphql_request(PRODUCTS_BY_TAG_QUERY, variables)
        
        products_data = data["products"]
        has_next_page = products_data["pageInfo"]["hasNextPage"]
        after = products_data["pageInfo"]["endCursor"]
        
        for edge in products_data["edges"]:
            products.append(edge["node"])
        
        page += 1
        time.sleep(0.5)
    
    print(f"  Fetched {len(products)} products with tag '{tag}'")
    return products


def fetch_collections() -> List[Dict]:
    """Fetch all collections."""
    print("Fetching collections...")
    collections = []
    has_next_page = True
    after = None
    page = 1
    
    while has_next_page:
        print(f"  Page {page}...")
        
        variables = {"first": 250, "after": after}
        data = make_graphql_request(COLLECTIONS_QUERY, variables)
        
        collections_data = data["collections"]
        has_next_page = collections_data["pageInfo"]["hasNextPage"]
        after = collections_data["pageInfo"]["endCursor"]
        
        for edge in collections_data["edges"]:
            collections.append(edge["node"])
        
        page += 1
        time.sleep(0.5)
    
    print(f"  Fetched {len(collections)} collections")
    return collections


def fetch_collection_products(collection_identifier: str) -> List[Dict]:
    """Fetch products from a collection by handle, ID, or title."""
    print(f"Fetching products from collection: {collection_identifier}")
    
    # First, get all collections to find the right one
    collections = fetch_collections()
    
    collection = None
    for coll in collections:
        if (coll["handle"] == collection_identifier or 
            coll["id"] == collection_identifier or 
            coll["title"] == collection_identifier):
            collection = coll
            break
    
    if not collection:
        raise SystemExit(f"Collection not found: {collection_identifier}")
    
    print(f"  Found collection: {collection['title']} (ID: {collection['id']})")
    
    products = []
    has_next_page = True
    after = None
    page = 1
    
    while has_next_page:
        print(f"  Page {page}...")
        
        variables = {"id": collection["id"], "first": 250, "after": after}
        data = make_graphql_request(COLLECTION_PRODUCTS_QUERY, variables)
        
        collection_data = data["collection"]
        if not collection_data:
            break
            
        products_data = collection_data["products"]
        has_next_page = products_data["pageInfo"]["hasNextPage"]
        after = products_data["pageInfo"]["endCursor"]
        
        for edge in products_data["edges"]:
            products.append(edge["node"])
        
        page += 1
        time.sleep(0.5)
    
    print(f"  Fetched {len(products)} products from collection")
    return products


def add_language_markers(products: List[Dict]) -> List[Dict]:
    """Add language markers to products."""
    products_with_lang = []
    
    for product in products:
        product_with_lang = product.copy()
        product_with_lang[f"title_{ORIGINAL_LANGUAGE}"] = product["title"]
        product_with_lang[f"descriptionHtml_{ORIGINAL_LANGUAGE}"] = product["descriptionHtml"]
        product_with_lang[f"title_{TARGET_LANGUAGE}"] = ""
        product_with_lang[f"descriptionHtml_{TARGET_LANGUAGE}"] = ""
        products_with_lang.append(product_with_lang)
    
    return products_with_lang


def ensure_output_dir(output_dir: str) -> Path:
    """Ensure output directory exists."""
    path = Path(output_dir)
    path.mkdir(parents=True, exist_ok=True)
    return path


def write_json(file_path: Path, data: List[Dict]) -> None:
    """Write data to JSON file."""
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def write_csv(file_path: Path, data: List[Dict]) -> None:
    """Write data to CSV file."""
    if not data:
        return
    
    import csv
    with open(file_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)


def write_xlsx(file_path: Path, data: List[Dict]) -> None:
    """Write data to XLSX file."""
    if not data:
        return
    
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Write headers
    headers = list(data[0].keys())
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Write data
    for row, product in enumerate(data, 2):
        for col, value in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=product.get(value, ""))
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, len(data) + 2):
            cell_value = ws[f"{column_letter}{row}"].value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    wb.save(file_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Fetch Shopify products with various methods",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Fetch all products
  python fetch_products.py all
  
  # Fetch single product by handle
  python fetch_products.py single --handle my-product
  
  # Fetch single product by ID
  python fetch_products.py single --id 123456789
  
  # Fetch products by tag
  python fetch_products.py tag --name featured
  
  # Fetch products from collection by handle
  python fetch_products.py collection --handle my-collection
  
  # Fetch products from collection by title
  python fetch_products.py collection --title "My Collection"
        """
    )
    
    subparsers = parser.add_subparsers(dest='command', help='Fetch method')
    
    # All products command
    all_parser = subparsers.add_parser('all', help='Fetch all products')
    
    # Single product command
    single_parser = subparsers.add_parser('single', help='Fetch single product')
    single_group = single_parser.add_mutually_exclusive_group(required=True)
    single_group.add_argument('--handle', help='Product handle')
    single_group.add_argument('--id', help='Product ID')
    
    # Tag command
    tag_parser = subparsers.add_parser('tag', help='Fetch products by tag')
    tag_parser.add_argument('--name', required=True, help='Tag name')
    
    # Collection command
    collection_parser = subparsers.add_parser('collection', help='Fetch products from collection')
    collection_group = collection_parser.add_mutually_exclusive_group(required=True)
    collection_group.add_argument('--handle', help='Collection handle')
    collection_group.add_argument('--title', help='Collection title')
    collection_group.add_argument('--id', help='Collection ID')
    
    # Common arguments
    parser.add_argument('--output-dir', default=OUTPUT_DIR, help='Output directory')
    parser.add_argument('--no-csv', action='store_true', help='Skip CSV export')
    parser.add_argument('--no-xlsx', action='store_true', help='Skip XLSX export')
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        sys.exit(1)
    
    validate_environment()
    output_dir = ensure_output_dir(args.output_dir)
    
    # Fetch products based on command
    products = []
    
    if args.command == 'all':
        products = fetch_all_products()
        filename_prefix = "products"
        
    elif args.command == 'single':
        if args.handle:
            product = fetch_product_by_handle(args.handle)
            if product:
                products = [product]
                filename_prefix = f"single_product_{args.handle}"
            else:
                print(f"Product not found: {args.handle}")
                sys.exit(1)
        elif args.id:
            product = fetch_product_by_id(args.id)
            if product:
                products = [product]
                filename_prefix = f"single_product_{args.id.replace('gid://shopify/Product/', '')}"
            else:
                print(f"Product not found: {args.id}")
                sys.exit(1)
                
    elif args.command == 'tag':
        products = fetch_products_by_tag(args.name)
        safe_name = args.name.replace(' ', '_').replace('/', '_')
        filename_prefix = f"products_tag_{safe_name}"
        
    elif args.command == 'collection':
        if args.handle:
            products = fetch_collection_products(args.handle)
            safe_name = args.handle.replace(' ', '_').replace('/', '_')
            filename_prefix = f"collection_{safe_name}"
        elif args.title:
            products = fetch_collection_products(args.title)
            safe_name = args.title.replace(' ', '_').replace('/', '_')
            filename_prefix = f"collection_{safe_name}"
        elif args.id:
            products = fetch_collection_products(args.id)
            filename_prefix = f"collection_{args.id.replace('gid://shopify/Collection/', '')}"
    
    if not products:
        print("No products found.")
        sys.exit(1)
    
    # Add language markers
    products_with_lang = add_language_markers(products)
    
    # Export files
    print(f"\nExporting {len(products)} products...")
    
    # Raw JSON
    raw_file = output_dir / f"{filename_prefix}_raw.json"
    write_json(raw_file, products)
    print(f"  - {raw_file}")
    
    # With language markers
    lang_file = output_dir / f"{filename_prefix}_with_lang.json"
    write_json(lang_file, products_with_lang)
    print(f"  - {lang_file}")
    
    # CSV (unless disabled)
    if not args.no_csv:
        csv_file = output_dir / f"{filename_prefix}_with_lang.csv"
        write_csv(csv_file, products_with_lang)
        print(f"  - {csv_file}")
    
    # XLSX (unless disabled)
    if not args.no_xlsx:
        xlsx_file = output_dir / f"{filename_prefix}_with_lang.xlsx"
        write_xlsx(xlsx_file, products_with_lang)
        print(f"  - {xlsx_file}")
    
    print(f"\n✅ Export complete! {len(products)} products exported.")


if __name__ == "__main__":
    main()